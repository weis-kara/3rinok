# Import `load_workbook` module from `openpyxl`
from openpyxl import load_workbook
import csv

dict_rinok = {'Всего': 40,
              'в том числе: универсальных': 44,
              'специализированных (по продаже продуктов питания)': 42,
              'Специализированный (по продаже радио- и электро-бытовой техники)': 49,
              'Специализированный (по продаже строительных материалов)': 45,
              'Специализированный (вещевой)': 41,
              'Специализированный сельскохозяйственный': 43,
              'Специализированный сельскохозяйственный кооперативный': 47,
              'Специализированный (прочие)': 46}

# Открываем файл Excel
workbook = load_workbook('input/3-рынок 3 кв 2023.xlsx')
value_rinok = dict_rinok['Всего']
print(value_rinok)
# Создаем CSV-файл для записи данных
csv_filename = 'output/output.csv'
csv_file = open(csv_filename, 'w', newline='')
csv_writer = csv.writer(csv_file, delimiter=';')

# Перебираем листы в файле Excel
for sheet_name in workbook.sheetnames:
    # Проверяем критерии выборки листа по его имени
    if not sheet_name.startswith('960') and not sheet_name.startswith('96600') and sheet_name.endswith('000'):
        sheet = workbook[sheet_name]
        # Перебираем строки в таблице
        for row in sheet.iter_rows(min_row=7, max_col=3, values_only=True):
            # Пропускаем пустые строки
            # if all(cell is None for cell in row):
            #    continue
            # Записываем название листа в начале строки
            row_data = [sheet_name]
            # Добавляем данные из строки таблицы
            row_data.extend(row)
            # приводим значения в столбцах к нужному формату
            row_data[0] = int(int(row_data[0])/1000) # октмо
            row_data[1] = " ".join(row_data[1].split()) # удаляем лишние пробелы в строках тип рынков
            row_data[1] = dict_rinok.get(row_data[1], 0) # во втором столбце заменяем тип рынка на его код
            if row_data[2] and row_data[3] is not None and row_data[1] > 0:
                # Записываем строку данных в CSV-файл
                csv_writer.writerow(row_data)

# Закрываем CSV-файл
csv_file.close()
print(f'Данные из Excel сохранены в {csv_filename}')

workbook.close()
